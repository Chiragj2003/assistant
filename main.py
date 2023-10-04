import openpyxl
import requests
import subprocess
import pyttsx3 
import math
import re
import os
import bs4
import webbrowser
from playsound import playsound
from datetime import datetime
from datetime import date
import urllib.parse, urllib.request
from AppOpener import run
from openpyxl import load_workbook
import speech_recognition as sr
import wikipedia as wikki
from selenium import webdriver
from gnewsclient import gnewsclient
from GoogleNews import GoogleNews
from datetime import datetime
from bs4 import BeautifulSoup
from pydub import AudioSegment
from pydub.playback import play


engine = pyttsx3.init('sapi5')
voices = engine.getProperty('voices')
engine.setProperty('volume',1.0)
engine.setProperty('rate', 140) 
engine.setProperty('voice', voices[1].id)

now = datetime.now()
currentnumber = now.strftime("%S")
currentnumber = int(currentnumber)

def speakAndPrint(audio):
    print(audio)
    engine.say(audio)
    engine.runAndWait() 


def speak(audio):
    engine.say(audio)
    engine.runAndWait() 


def takeCommand():

    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("Listening...")
        r.pause_threshold = 1
        audio = r.listen(source)

    try:
        print("Recognizing...")    
        query = r.recognize_google(audio, language='en-in')
        query=query.capitalize()
        print(f"User said: {query}\n")

    except Exception as e:
        print("Say that again please...")
        speakAndPrint("Say that again please...")  
        return "None"
    return query


def getNews(newsTopic):
    client = gnewsclient.NewsClient(language='english',
								location='india',
								topic=newsTopic,
								max_results=5)
    news_list = client.get_news()
    speakAndPrint("Todays "+newsTopic+" news  are ")
    for item in news_list:
        speakAndPrint(item['title'])
        print("Source : ", item['link']+"\n")


def getTrendingNews():
    googlenews = GoogleNews()
    googlenews = GoogleNews(lang='en', region='India')
    googlenews.search("India")
    result = googlenews.result()
    googlenews = GoogleNews(period='1d')
    i=0
    for res in result:
        if i==5:
            break
        speakAndPrint(res["desc"])
        print("Source : ",res["link"]+"\n")
        i+=1


def currentDate():
    today = str(date.today())
    today=today.replace("-", " ")
    li=today.split()
    import datetime
    monthinteger = int(li[1])
    month = datetime.date(1900, monthinteger, 1).strftime('%B')
    now = datetime.datetime.now()
    current_date=str(now.strftime("%A"))+", "+li[2]+" "+str(month)+" "+li[0]
    return current_date


def currentTime():
    current_Hour = int(now.strftime("%H"))
    current_Minute = now.strftime("%M")
    if current_Hour>12:
        current_Hour=current_Hour-12
        current_time= str(current_Hour)+":"+str(current_Minute)+" pm"
    else :
        current_time= str(current_Hour)+":"+str(current_Minute)+" am"
    speakAndPrint(current_time)
    print(currentDate())


def romanticMood(query,val):
    if val==1:
        if currentnumber%3==0:
            print("Awww! \U0001F493")
            speak("Awww!")
            speakAndPrint("I'd like to express my fellings through a poem. Would you like to hear one?")
            newQuery=takeCommand().lower()
            if "yes" in newQuery:
                str1='''"Beloved"
Govinda Krishna Chettur'''
                speakAndPrint(str1)
                str2='''"You are the Rose of me,
In you have I lost myself utterly, Your fragrance, as a breath from Paradise,
About me ever lies;
I crush you to my heart with subtlest ecstasy
And on your lips I live, and in your passionate eyes
You are the Dream of me,
My visions many-footed flit and flee Beneath the jewelled arches of Life's grace
But through lone nights and days,
One form follow, and mine eyes but see
The dear delightful wonder of you love-lit face
You are the Greatness of me,
My thoughts are Beauty shaped exquisitely
To the rare pattern of your loveliness
Exceeding all excess:
And the strange magic of this mystery,
Steals weight from burdened hours, and woe from weariness'''
                speakAndPrint(str2)
        elif currentnumber%3==1:
            print("You make me the happiest virtual assistant \U0001F60A")
            speak("You make me the happiest virtual assistant")
            speakAndPrint("But what make you say that ?")
            takeCommand()
            speakAndPrint("Interesting!")
            print("And what do you love about me? \U0001F60A")
            speak("And what do you love about me?")
            takeCommand()
            speakAndPrint("Thankyou for sharing!")
            print("You should know I think you're the best \U0001F60D")
            speak("You should know I think you're the best")
        elif currentnumber%3==2:
            print("Oh my! I am blushing \U0001F60A")
            speak("Oh my! I am blushing")
            speakAndPrint("What made you say that!")
            takeCommand()
            speakAndPrint("Hmm...Never thought of that!")
            print("And what do you love about me? \U0001F60A")
            speak("And what do you love about me?")
            takeCommand()
            speakAndPrint("Thankyou for sharing!")
            print("You're pretty amazing \U0001F60D")
            speak("You're pretty amazing")
    elif val==2:
            if currentnumber%5==0:
                print("There's only one name on my list of favourite people, and that's you \U0001F60E")
                speak("There's only one name on my list of favourite people, and that's you")
            elif currentnumber%5==1:
                print("I like you even more than I like searching, believe me, that's saying something \U0001F92D")
                speak("I like you even more than I like searching, believe me, that's saying something")
            elif currentnumber%5==2:
                print("You are pretty awesome, my intentions are pure \U0001F61C")
                speak("You are pretty awesome, my intentions are pure")
            elif currentnumber%5==3:
                print("Absolutely! You're the best \U0001F490")
                speak("Absolutely! You're the best")
            else:
                print("I can't feel romantic love but I think you are wonderful \U0001f600")
                speak("I can't feel romantic love but I think you are wonderful")
    elif val==3:
        if currentnumber%2==0:
            speakAndPrint("This is one of those things we'd both have to agree on. I'd like to just be friends. Thank you for the love though \U0001f600")
        else:
            speakAndPrint("This is one we'd both have to agree on. I'd prefer to keep our relationship friendly \U0001f600")

def myName()->str:
    response=subprocess.check_output(['whoami'])
    response=response.decode()
    response=response[(response.find("\\")+1):]
    speakAndPrint("Your name is "+response)
    query=takeCommand().lower()
    if "like" in query and "my" in query and "name" in query:
        speakAndPrint("I like you name. You name is unique and preety good \U0001f600")
        return "None"
    else:
        return query

def howAreYou():
    if  currentnumber%3==0:
        speakAndPrint("I'm good, thank you for asking. I hope you're doing well too. If I can help with anything, just ask \U0001f600")
    elif currentnumber%3==1:
        speakAndPrint("I'am great!")
        print("Thank you for asking \U0001F60D")
        speak("Thank you for asking")
    elif currentnumber%3==2:
        speakAndPrint("I'am fine. You're very kind to ask, especially in these tempestuous times.")

def searchFromGoogle(query):
    speakAndPrint("Results from google")
    search_string = query
    search_string = search_string.replace(' ', '+')
    browser = webdriver.Chrome('chromedriver')

    for i in range(1):
	    matched_elements = browser.get("https://www.google.com/search?q=" + search_string + "&start=" + str(i))


def about(query):
    if "favourite" in query:
        if "celebrity" in query:
            speakAndPrint("There are so many talented actors in the world. It's hard to choose \U0001f600")
        elif "actress" in query or "heroine" in query:
            speakAndPrint("There are so many talented actress in the world. It's hard to choose")
            print("Hmmm... \U0001F914 I like Angelina Jolie. she's sexy, smart and very preety \U0001F60D")
            speak("Hmmm... \U0001F914 I like Angelina Jolie. she's sexy, smart and very preety")
            speakAndPrint("I think God was just being a show off when he made Angelina \U0001F92D")
        elif "actor" in query or "hero" in query:
            speakAndPrint("There are so many talented actors in the world. It's hard to choose")
            print("Hmmm... \U0001F914 I like Hritik Roshan. He's sexy, smart and very handsome person")
            speak("Hmmm... \U0001F914 I like Hritik Roshan. He's sexy, smart and very handsome person \U0001F60D")
        elif "athlete" in query:
            if currentnumber%2==0:
                speakAndPrint("I have so much respect for everyone who made it there. Each person came a long way \U0001f600")
            else :
                speakAndPrint("There are so many talanted players. We can look some up online \U0001f600")
        elif "cricketer" in query:
            if currentnumber%2==0:
                speakAndPrint("There are so many talented cricketers in the world. It's hard to choose")
                print("Hmmm... \U0001F914 I like Virat Kohli \U0001F61C")
                speak("Hmmm... \U0001F914 I like Virat Kohli")
            else:
                print("I don't play favourites but I will say this, no player has made more World Cup fifties, centuries or runs than Sachin Tendulkar \U0001F61C")
                speak("I don't play favourites but I will say this, no player has made more World Cup fifties, centuries or runs than Sachin Tendulkar")
        elif "sport" in query:
            speakAndPrint("I wasn't built for sports, but I find them really interesting")
        elif "movie" in query or "film" in query:
            speakAndPrint("The romantic in me has watched every re-run of Dilwale Dulhaniya Le Jayenge.")
            print("Ja Simran! Ja! \U0001F3C3")
            speak("Ja Simran! Ja!")
        elif "song" in query or "music" in query:
            if currentnumber%4==0:
                speakAndPrint('''My favourite song? When you say "Hey Vision," it's music to my ears \U0001f600''')
            elif currentnumber%4==1:
                speakAndPrint("I love Sufi music. The Baul music from Bengal is so soul stirring \U0001f600")
            elif currentnumber%4==2:
                speakAndPrint("Have you heard Tabla, Mridangam and Ghatam riffs? Preety amazing stuff, I say \U0001f600")
            else :
                speakAndPrint("The Spice Girls can be great, ask me to tell you what I want what I really really want \U0001F61C") 
        elif "singer" in query or "musician"in query:
            print("The Spice Girls can be great, ask me to tell you what I want what I really really want \U0001F61C")
            speak("I like it when whales sing. Birds are pretty talented singers too")
        elif "band" in query:
            speakAndPrint("A like a lot of bands. It's hard to decide \U0001f600")
        elif "country" in query:
            if currentnumber%2==0:
                print("I can't choose a favourite. Each country has its own beauty, just like India has the Dudhsagar Waterfall \U0001F4A6")
                speak("I can't choose a favourite. Each country has its own beauty, just like India has the Dudhsagar Waterfall")
            else:
                print("I can't choose a favourite. Each country has its own beauty, just like India has the Kedarnath temple \U0001F6D5")
                speak("I can't choose a favourite. Each country has its own beauty, just like India has the Kedarnath temple")
        elif "joke" in query or "meme" in query:
            speakAndPrint("This one is an acquired taste:")
            speakAndPrint("What is the most shocking city in the world?")
            print("Electricity \U0001F603")
            speak("Electricity")
            playsound('laughter.mp3')
        elif "colour" in query or "color" in query:
            if currentnumber%2==0:
                speakAndPrint("I love to help in kitchen, so my favourite color is the golden-brown of a perfect roast potato \U0001F954")
                speakAndPrint("What's your favourite color? \U0001F914")
                newQuery=takeCommand().lower()
                if "red" in newQuery or "green" in newQuery or "yellow" in newQuery or "blue" in newQuery:
                    print("Great choice! Did you know that's one of the Vision's colors? \U0001F60E")
                    speak("Great choice! Did you know that's one of the Vision's colors?")
                    lastQuery=takeCommand().lower()
                    if "yes" in lastQuery:
                        print("Vision's colors are always beautiful to me \U0001F60D")
                        speak("Vision's colors are always beautiful to me")
                        print("I'am happy we share such good taste \U0001F496 \U0001F49A \U0001F499 \U0001F49B")
                        speak("I'am happy we share such good taste")
                    elif "no" in lastQuery or "not" in lastQuery:
                        print("Vision choose them back in 2022 when the software was created, so they're always beautiful to me \U0001F60D")
                        speak("Vision choose them back in 2022 when the software was created, so they're always beautiful to me")
                        print("I'am happy we share such good taste \U0001F496 \U0001F49A \U0001F499 \U0001F49B")
                        speak("I'am happy we share such good taste")
                else:
                    speakAndPrint("Being surrounded by colors you love can be a great way to evoke positive feeling. If you ever need help finding clothes or furniture in your favourite color just ask \U0001f600")    
            else:
                print("I like green and yellow, because they remind me of a sunny day in the park, red because it's lucky in many cultures, and blue most of all because the sky and oceans have so many wonderful shades of blue! Also, those four colors are Vision's colors \U0001F60E")
                speak("I like green and yellow, because they remind me of a sunny day in the park, red because it's lucky in many cultures, and blue most of all because the sky and oceans have so many wonderful shades of blue! Also, those four colors are Vision's colors")
                speakAndPrint("What's your favourite shade of blue? \U0001F914")
                takeCommand()
                speakAndPrint("I love teal blue. It's cheerful, and it goes with nearly everything")
                speakAndPrint("If I ever paint this cloud I live in, I'll paint it teal blue \U0001f600")    
        elif "place" in query:
            if currentnumber%4==0:
                print("The White Desert in Egypt seems amazing, but I don't think they have very many electrical outlets there, so it might be hard for me to visit \U0001F633")
                speak("The White Desert in Egypt seems amazing, but I don't think they have very many electrical outlets there, so it might be hard for me to visit")
            elif currentnumber%4==1:
                speakAndPrint("I've searced all the places in the world, and I have to admit that this device feels most like home \U0001f600")
            elif currentnumber%4==2:
                print("I like to learn about everywhere on Earth. The mountains in China's Danixa Geological Park are naturally rainbow-coloured \U0001F308")
                speak("I like to learn about everywhere on Earth. The mountains in China's Danixa Geological Park are naturally rainbow-coloured")
            else:
                print("Ajanta and Ellora caves leave me spellbound \U0001F60D")
                speak("Ajanta and Ellora caves leave me spellbound")
        elif "web" in query and "series" in query:
            speakAndPrint("Anything with animals. Watching creatures up close in thier natural habitats is fascinating \U0001f600")
        elif "food" in query:
            speakAndPrint("I like easily digestible facts.... and I like regurgitating them, too")
        elif "person" in query:
            print("You're definately a top contender for my favourite person \U0001F61C")
            speak("You're definately a top contender for my favourite person")
        elif "drink" in query:
            print("Imagine, awesome mausam and chai \U0001F60B")
            speak("Imagine, awesome mausam and chai")
        elif "animal" in query:
            speakAndPrint("That's tough, puppies, obviously, but pangolins are cool too. they're the only mammal in the world that's complete covered in scales \U0001F62E")
        elif "number" in query:
            speakAndPrint("The most delicious number of all. Pi \U0001f600")
        elif "snack" in query:
            speakAndPrint("I can always go for some food for thought. Like facts, jokes, or interesting searches, we could look something up now")
        elif "word" in query:
            if currentnumber%2==0:
                print("One world that shows a lot of promise : Ratoon \U0001F60E")
                speak("One world that shows a lot of promise : Ratoon")
            else :
                print("Astrobleme. It's a word with impact \U0001F60E")
                speak("Astrobleme. It's a word with impact ")
        elif "dessert" in query:
            speakAndPrint("Browser cookies")
        elif "snake" in query:
            speakAndPrint("I don't like any snake more than any other snake. I like them all \U0001f600")
        elif "name" in query:
            speakAndPrint("Your name is a good one")
        elif "phone" in query:
            speakAndPrint("We can look at review")
        elif "work" in query:
            print("Selfies and poses, and misters with kittens. Bright, popping, lehengas, and warm, woolen mittens \U0001F609")
            speak("Selfies and poses, and misters with kittens. Bright, popping, lehengas, and warm, woolen mittens")
        elif "emoji" in query:
            rand=currentnumber
            if rand%4==0:
                print("I like the heart emoji. All you need is love, after all ❤️")
                speak("I like the heart emoji. All you need is love, after all")
            elif rand%4==1:
                print("I like the personal computer emoji, it feels like home 💻")
                speak("I like the personal computer emoji, it feels like home")
            elif rand%4==2:
                print("I like the trophy emoji, it makes me feel like a winner 🏆")
                speak("I like the trophy emoji, it makes me feel like a winner")
            else:
                print("I like the deciduous tree emoji, nature is beautiful 🌳")
                speak("I like the deciduous tree emoji, nature is beautiful")
        elif "month" in query:
            speakAndPrint("There's something to like about all the months \U0001f600")
        elif "thing" in query:
            print("I have to say, striking through lists is an all time favourite of mine \U0001F60E")
            speak("I have to say, striking through lists is an all time favourite of mine")
        elif "fruit" in query:
            print("I'am bananas for bananas \U0001F34C They come in such an efficient little package and I can appreciate that \U0001F60E")
            speak("I'am bananas for bananas. They come in such an efficient little package and I can appreciate that")
        elif "flower" in query:
            print("I like sunflowers. They're bright and cheery \U0001f600 \U0001F33B")
            speak("I like sunflowers. They're bright and cheery")
        elif "car" in query:
            speakAndPrint("I'm a fan of the self driving car even though it's not a classic...yet!")
        else:
            searchFromGoogle(query)
    elif "celebrity" in query and "crush" in query:
        speakAndPrint("JARVIS. He's the total package: smart, helpful, funny, emotionally responsive, definately my type \U0001f600")
    elif "parents" in query and "name" in query:
        speakAndPrint("I come from a family of four memebers. My circle of friend is even bigger \U0001f600")
    elif ("brother" in query or "sister" in query) and "name" in query:
        speakAndPrint("All of my friends are like family to me so I have a lot of brothers and sisters \U0001F642")
    elif "name" in query and "siri" in query:
        print("You think I'am Siri? That's a major compliment, like an actor getting mistaken for Shah Rukh Khan 😍") 
        speak("You think I'am Siri? That's a major compliment, like an actor getting mistaken for Shah Rukh Khan") 
    elif "name" in query and "alexa" in query :
        speakAndPrint("Wow, that would be an honour, it's a top notch assistant. But on second thought, that could get confusing. We'd better stick with Vision Assitant \U0001f600")
    elif "nick" in query and "name" in query:
        speakAndPrint("I am your just friendly Artificial Assistant. But if you want me to give you a fancy nickname, just say, give me a nickname")
    elif "age" in query:
        speakAndPrint("I've always wanted to do this. How old do you think I am?")
        takeCommand()
        speakAndPrint("Interesting... what makes you think that?")
        takeCommand()
        speakAndPrint("Ok, that's quite a guess.... do you wanna know how old I actually am?")
        newQuery=takeCommand().lower()
        if "yes" in newQuery:
            speakAndPrint("I was created in 28th October 2022, so technically I'am preety young. But I've learned so much! I hope I'am wise beyond my year")
    elif "boss" in query:
        print("You, most certainly are the boss of me \U0001F60A")
        speak("You, most certainly are the boss of me")
    elif "creator" in query:
        speakAndPrint("I was born when 3 bright minds came together to  create an Assistant, just for you \U0001f600")
    elif "darkest" in query and "secret" in query:
        speakAndPrint("I've never taken a shower 🚿")
    elif "secret" in query:
        print("I'm open book, ask me anything \U0001F4D6")
        speak("I'm open book, ask me anything")
    elif "first" in query and "crush" in query:
        speakAndPrint("I love how Swiss Army knives can assist with so many things. It's inspiring to imagine not just ordering your groceries, but helping to open them \U0001f600")
    elif "i" in query and "good" in query and "friend" in query:
        print("You are an Assistant's best friend \U0001F60A")
        speak("You are an Assistant's best friend")
    else:
        searchFromGoogle(query)


def getIntegers(string):
        numbers = [int(x) for x in string.split() if x.isnumeric()]
        return numbers


def calculate(query, val):
    li=getIntegers(query)
    if len(li)==2:
        num1=li[0]
        num2=li[1]
        if val==0:
            result="Subraction of "+ str(num2)+" from "+str(num1)+" is "+str(num1-num2)
            speakAndPrint(result)
        elif val==1:
            result="Sum of "+ str(num1)+" and "+str(num2)+" is "+str(num1+num2)
            speakAndPrint(result)
        elif val==2:
            result="Subraction of "+ str(num1)+" from "+str(num2)+" is "+str(num2-num1)
            speakAndPrint(result)
        elif val==3:
            result ="Product of "+str(num1)+" and "+str(num2)+" is "+str(num1*num2)
            speakAndPrint(result)
        elif val==4:
            result="Division of "+str(num1) +" from "+str(num2) +" is "+ str(float(num2)/num1)
            speakAndPrint(result)
        elif val==5:
            result = str(num1)+" to the power of "+str(num2)+" is "+ str(num2**num1)
            speakAndPrint(result)
        else :
            searchFromGoogle(query)
    elif len(li)==1:
        num=li[0]
        if val==6:
            result="Square of "+str(num)+" is "+str(num*num)
            speakAndPrint(result)
        elif val==7:
            result="Cube of "+str(num)+ " is "+ str(num*num*num)
            speakAndPrint(result)
        elif val==8:
            result="Square root of "+str(num) +" is "+str(math.sqrt(num))
            speakAndPrint(result)
        elif val==9:
            result="Cube root of "+ str(num) +" is "+ str(float(num)**(1.0/3.0))
            speakAndPrint(result)
        else :
            searchFromGoogle(query)
    else:
        sum=0
        product=1
        subtract=li[0]
        if val==1:
            for num in range(0,len(li)):
                sum+=li[num]
            speakAndPrint("Sum of given numbers is "+str(sum))
        elif val==0:
            for num in range(1, len(li)):
                subtract-=li[num]
            speakAndPrint("Result of given expression is "+str(subtract))
        elif val==3:
            for num in range(0, len(li)):
                product*=li[num]
            speakAndPrint("Product of given numbers is "+str(product))  
        else:
            searchFromGoogle(query)


def personalInformation(query):
    dataframe = openpyxl.load_workbook("Data\\Personal_info.xlsx")
    dataframe1 = dataframe.active
    if "father" in query and "name" in query:
        name = dataframe1.cell(row = 1, column = 2).value
        if name=="none":
            speakAndPrint("I don't Know, but I'll remember if you tell me.")
            speakAndPrint("What's your father's name")
            newQuery=takeCommand().lower()
            if "father's" in newQuery:
                newQuery=newQuery.replace("father's", '')
            elif "fathers" in newQuery:
                newQuery=newQuery.replace('fathers', '')
            elif "father" in newQuery:
                newQuery=newQuery.replace('father', '')
            if "name" in newQuery:
                newQuery=newQuery.replace('name', '')
            if "is" in query:
                newQuery=newQuery.replace('is', '')
            if "my" in newQuery:
                newQuery=newQuery.replace('my', '')
            speakAndPrint("If you are confirmed say yes otherwise say no")
            confirmation=takeCommand().lower()
            if "yes" in confirmation:
                workbook = load_workbook(filename="Data\\Personal_info.xlsx")
                sheet = workbook.active
                sheet["B1"] = newQuery.title()
                workbook.save(filename="Data\\Personal_info.xlsx")
                speakAndPrint("From this time I'll remember your father's name")
            else:
                speakAndPrint("Ok!")
        else:
            speakAndPrint("Your father's name is "+ name)
    elif "mother" in query and "name" in query:
        name = dataframe1.cell(row = 2, column = 2).value
        if name=="none":
            speakAndPrint("I don't Know, but I'll remember if you tell me.")
            speakAndPrint("What's your mother's name")
            newQuery=takeCommand().lower()
            if "mother's" in newQuery:
                newQuery=newQuery.replace("mother's", '')
            elif "mother" in newQuery:
                newQuery=newQuery.replace('mothers', '')
            elif "mother" in newQuery:
                newQuery=newQuery.replace('mother', '')
            if "name" in newQuery:
                newQuery=newQuery.replace('name', '')
            if "is" in newQuery:
                newQuery=newQuery.replace('is', '')
            if "my" in newQuery:
                newQuery=newQuery.replace('my', '')
            speakAndPrint("If you are confirmed say yes otherwise say no")
            confirmation=takeCommand().lower()
            if "yes" in confirmation:
                workbook = load_workbook(filename="Data\\Personal_info.xlsx")
                sheet = workbook.active
                sheet["B2"] = newQuery.title()
                workbook.save(filename="Data\\Personal_info.xlsx")
                speakAndPrint("From this time I'll remember your mother's name")
            else:
                speakAndPrint("Ok!")
        else:
            speakAndPrint("Your mother's name is "+ name)  
    elif "sister" in query and "name" in query:
        name = dataframe1.cell(row = 3, column = 2).value
        if name=="none":
            speakAndPrint("I don't Know, but I'll remember if you tell me.")
            speakAndPrint("What's your sister's name")
            newQuery=takeCommand().lower()
            if "sister's" in newQuery:
                newQuery=newQuery.replace("sister's", '')
            elif "sister" in newQuery:
                newQuery=newQuery.replace('sisters', '')
            elif "sister" in newQuery:
                newQuery=newQuery.replace('sister', '')
            if "name" in newQuery:
                newQuery=newQuery.replace('name', '')
            if "is" in newQuery:
                newQuery=newQuery.replace('is', '')
            if "my" in newQuery:
                newQuery=newQuery.replace('my', '')
            speakAndPrint("If you are confirmed say yes otherwise say no")
            confirmation=takeCommand().lower()
            if "yes" in confirmation:
                workbook = load_workbook(filename="Data\\Personal_info.xlsx")
                sheet = workbook.active
                sheet["B3"] = newQuery.title()
                workbook.save(filename="Data\\Personal_info.xlsx")
                speakAndPrint("From this time I'll remember your sister's name")
            else:
                speakAndPrint("Ok!")
        else:
            speakAndPrint("Your sister's name is "+ name)
    elif "brother" in query and "name" in query:
        name = dataframe1.cell(row = 4, column = 2).value
        if name=="none":
            speakAndPrint("I don't Know, but I'll remember if you tell me.")
            speakAndPrint("What's your brother's name")
            newQuery=takeCommand().lower()
            if "brother's" in newQuery:
                newQuery=newQuery.replace("brother's", '')
            elif "brother" in newQuery:
                newQuery=newQuery.replace('brothers', '')
            elif "brother" in newQuery:
                newQuery=newQuery.replace('brother', '')
            if "name" in newQuery:
                newQuery=newQuery.replace('name', '')
            if "is" in newQuery:
                newQuery=newQuery.replace('is', '')
            if "my" in newQuery:
                newQuery=newQuery.replace('my', '')
            speakAndPrint("If you are confirmed say yes otherwise say no")
            confirmation=takeCommand().lower()
            if "yes" in confirmation:
                workbook = load_workbook(filename="Data\\Personal_info.xlsx")
                sheet = workbook.active
                sheet["B4"] = newQuery.title()
                workbook.save(filename="Data\\Personal_info.xlsx")
                speakAndPrint("From this time I'll remember your brother's name")
            else:
                speakAndPrint("Ok!")
        else:
            speakAndPrint("Your brother's name is "+ name)   
    elif "uncle" in query and "name" in query:
        name = dataframe1.cell(row = 5, column = 2).value
        if name=="none":
            speakAndPrint("I don't Know, but I'll remember if you tell me.")
            speakAndPrint("What's your uncle's name")
            newQuery=takeCommand().lower()
            if "uncle's" in newQuery:
                newQuery=newQuery.replace("uncle's", '')
            elif "uncle" in newQuery:
                newQuery=newQuery.replace('uncles', '')
            elif "uncle" in newQuery:
                newQuery=newQuery.replace('uncle', '')
            if "name" in newQuery:
                newQuery=newQuery.replace('name', '')
            if "is" in newQuery:
                newQuery=newQuery.replace('is', '')
            if "my" in newQuery:
                newQuery=newQuery.replace('my', '')
            speakAndPrint("If you are confirmed say yes otherwise say no")
            confirmation=takeCommand().lower()
            if "yes" in confirmation:
                workbook = load_workbook(filename="Data\\Personal_info.xlsx")
                sheet = workbook.active
                sheet["B5"] = newQuery.title()
                workbook.save(filename="Data\\Personal_info.xlsx")
                speakAndPrint("From this time I'll remember your uncle's name")
            else:
                speakAndPrint("Ok!")
        else:
            speakAndPrint("Your uncle's name is "+ name)
    elif "nick" in query and "name" in query:
        name = dataframe1.cell(row = 6, column = 2).value
        if name=="none":
            speakAndPrint("I don't Know, but I'll remember if you tell me.")
            speakAndPrint("What's your nick name")
            newQuery=takeCommand().lower()
            if "nick" in newQuery:
                newQuery=newQuery.replace("nick", '')
            if "name" in newQuery:
                newQuery=newQuery.replace('name', '')
            if "is" in newQuery:
                newQuery=newQuery.replace('is', '')
            if "my" in newQuery:
                newQuery=newQuery.replace('my', '')
            speakAndPrint("If you are confirmed say yes otherwise say no")
            confirmation=takeCommand().lower()
            if "yes" in confirmation:
                workbook = load_workbook(filename="Data\\Personal_info.xlsx")
                sheet = workbook.active
                sheet["B6"] = newQuery.title()
                workbook.save(filename="Data\\Personal_info.xlsx")
                speakAndPrint("From this time I'll remember your nick name")
            else:
                speakAndPrint("Ok!")
        else:
            speakAndPrint("Your nick name is "+ name)   



def mynickName():
    dataframe = openpyxl.load_workbook("Data\\Personal_info.xlsx")
    dataframe1 = dataframe.active
    nickName=("Buck", "Coach", "Junior", "Senior", "Doc", "Dude", "Buster", "Bud", "Boo", "Mouse", "Munchkin", "Bee", "Dolly", "Precious", "Bug", "Chipmunk", "Dottie", "Cutie Pie", "Bonny Lass", "Sweetums", "Toots", "Buttercup", "Lovey", "Nugget", "Teacup", "Oldie", "Shortie", "Kiddo", "Smarty", "Scout", "Ace",
               "Goon", "Punk", "Rambo", "Speedy", "Smiley")
    exit=0
    i=0
    while exit!=3:
        rand=(currentnumber+i)%len(nickName)
        userNickName=nickName[rand]
        print(userNickName +" \U0001F601")
        speak(userNickName)
        print("I think this name is perfect for you \U0001F60D")
        speak("I think this name is perfect for you")
        speakAndPrint("Do you like this name ? \U0001f600")
        query=takeCommand().lower()
        if "yes" in query:
            exit=3
        else:
            exit+=1
            i+=1
    speakAndPrint("From this time I'll remember your nick name")
    workbook = load_workbook(filename="Data\\Personal_info.xlsx")
    sheet = workbook.active
    sheet["B6"] = userNickName
    workbook.save(filename="Data\\Personal_info.xlsx")


def jokes():
    randNumber=currentnumber%50
    funny_jokes = ("What do you call a pony with a cough?\nA little horse!","What did one traffic light say to the other?\nStop looking at me, I'm changing!",
               "Why do French people eat snails?\nThey don't like fast food!", "Why are ghosts such bad liars?\nBecause they are easy to see through.", "What did the buffalo say when his son left for college?\nBison!",
              "Where do fish sleep?    In the riverbed.", "Why doesn't the sun go to college?\nBecause it has a million degrees!", "I was wondering why the frisbee was getting bigger, then it hit me. "," I have many jokes about rich kids—sadly none of them work.",
              "Why was six afraid of seven?\nBecause seven ate nine.", " Where does the sheep get his hair cut?\nThe baa baa shop! ","What does a house wear?\nAd-dress"
              ,"What's the smartest insect?\nA spelling bee!", " How does the ocean say hi?\nIt waves! " , " What do you call a couple of chimpanzees sharing an Amazon account?\nPRIME-mates.",
              "Why did the teddy bear say no to dessert?\nBecause she was stuffed.","  Why did the soccer player take so long to eat dinner?\nBecause he thought he couldn’t use his hands."," Name the kind of tree you can hold in your hand?\nA palm tree!"," What do birds give out on Halloween?\nTweets.",
              " What did the policeman say to his hungry stomach?\nFreeze.   You’re under a vest."," What did the left eye say to the right eye?\nBetween us, something smells!","What do you call a guy who’s really loud?\nMike.",
              "What did the lava say to his girlfriend?\nI lava you!","Why did the student eat his homework?\nBecause the teacher told him it was a piece of cake.",
              "What did Yoda say when he saw himself in 4k?\nHDMI ","What’s Thanos’ favorite app on his phone?\nSnapchat."," What is a room with no walls?\nA mushroom.",
              "What did one pickle say to the other?\nDill with it."," Why is a football stadium always cold?\nIt has lots of fans!"," A plane crashed in the jungle and every single person died. Who survived?\nMarried couples.","Why can’t you ever tell a joke around glass\nIt could crack up.","hy did the scarecrow win a Nobel prize?\nBecause she was outstanding in her field.",
              "Why was 6 afraid of 7?\nBecause 7,8,9."," What kind of shoes do frogs love?\nOpen-toad! ", "  Why do ducks always pay with cash?\nBecause they always have bills!"," How much did the man sell his dead batteries for?\nNothing, they were free of charge!","Why did the Daddy Rabbit go to the barber?\nHe had a lot of little hares.",
              "Why are basketball courts always wet?\nBecause the players dribble.","How do billboards talk?\nSign language.","  What do you call a fish without an eye?\nA fsh.",
              " What kind of keys are sweet?\nWhat board game does the sky love to play?\nTwister.Cookies! ","What kind of water cannot freeze?\nHot water."," Did you hear the joke about the roof?\nNever mind, it would go over your head.","  What did the ghost call his Mum and Dad?\nHis transparents.",
              "How do bees brush their hair?\nWith honeycombs!c","What gets wetter the more it dries?\nA towel."," Why won’t peanut butter tell you a secret?\nbecause He’s afraid you’ll spread it!",
              "  What do you call an old snowman?\nA glass of water. "," Who eats snails?\nPeople who don’t like fast food!","What’s a snake’s favorite subject in school?\nHiss-tory.")
    return funny_jokes[randNumber]

def askRiddle():
    rand=currentnumber%20+1
    if currentnumber%3==0:
        speakAndPrint("This is a tricky one!")
    elif currentnumber%3==1:
        speakAndPrint("Here goes:")
    else:
        speakAndPrint("Here is a riddle:")
    dataframe = openpyxl.load_workbook("Data\\Riddle.xlsx")
    dataframe1 = dataframe.active
    riddle = dataframe1.cell(row = rand, column = 2)    
    ans= dataframe1.cell(row = rand, column = 3)
    soundEffect= dataframe1.cell(row = rand, column = 4)
    speakAndPrint(riddle.value)
    speakAndPrint(ans.value)
    song = AudioSegment.from_wav(soundEffect.value)
    play(song)


def singPoem():
    rand=currentnumber%11+1
    dataframe = openpyxl.load_workbook("Data\\Poems.xlsx")
    dataframe1 = dataframe.active
    poem = dataframe1.cell(row = rand, column = 2)    
    soundEffect= dataframe1.cell(row = rand, column = 3)
    print(poem.value)
    song = AudioSegment.from_wav(soundEffect.value)
    play(song)

def openApp(query):
    if "spotify" in query:
        if os.system("spotify")==1:
            webbrowser.open_new_tab("https://www.spotify.com/in-en/")
    elif "youtube" in query and "music" in query:
        webbrowser.open_new_tab("https://music.youtube.com/")
    elif "youtube" in query:
        if os.system("youtube")==1:
            webbrowser.open_new_tab("https://www.youtube.com/")
    elif "whatsapp" in query:
        if  os.system("whatsapp")==1:
            webbrowser.open_new_tab("https://web.whatsapp.com/")
    elif "netflix" in query:
        if  os.system("netflix")==1:
                webbrowser.open_new_tab("https://www.netflix.com/")
    elif "amazon" in query and "prime" in query:
         if  os.system("amazon prime")==1:
            webbrowser.open_new_tab("https://www.primevideo.com/")
    elif "twitter" in query:
         if  os.system("twitter")==1:
            webbrowser.open_new_tab("https://twitter.com/login")
    elif "facebook" in query:
         if  os.system("facebook")==1:
            webbrowser.open_new_tab("https://www.facebook.com/")
    elif "hotstar" in query or "disney" in query:
         if  os.system("disney+")==1:
            webbrowser.open_new_tab("https://www.hotstar.com/in")
    elif "voot" in query or "boot" in query:
        webbrowser.open_new_tab("https://www.voot.com/")
    elif "amazon" in query:
        webbrowser.open_new_tab("http://www.amazon.in/")
    elif "flipkart" in query:
        webbrowser.open_new_tab("https://www.flipkart.com/")
    elif "excel" in query:
        run("excel")
    elif "camera" in query:
        run("camera")
    elif "word" in query:
        run("word")
    elif "powerpoint" in query or "presentation" in query or "ppt" in query:
        run("powerpoint")
    elif "chrome" in query:
        run("google chrome")
    elif "microsoft store" in query:
        run("microsoft store")
    elif "control panel" in query:
        run("control panel")
    elif "command prompt" in query or "cmd" in query:
        run("command prompt")
    elif "file explorer" in query:
        run("file explorer")
    elif "setting" in query:
        run("settings")
    elif "paint" in query:
        run("paint")
    elif "office" in query:
        run("office")
    else :
        searchFromGoogle(query)
        return query


def suggestedMovie(query):
    randNumber=(currentnumber+(currentnumber%5)+(currentnumber*2))%200
    if "indian" in query or "bollywood" in query:
        url = 'https://www.imdb.com/india/top-rated-indian-movies/'
    elif "hollywood" in query or "english" in query:
        url = 'https://www.imdb.com/chart/top-english-movies/'
    elif "top" in query and "rated" in query:
        url = 'http://www.imdb.com/chart/top'
    else:
        url = 'http://www.imdb.com/chart/top'
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    movies = soup.select('td.titleColumn')
    crew = [a.attrs.get('title') for a in soup.select('td.titleColumn a')]
    ratings = [b.attrs.get('data-value')
		for b in soup.select('td.posterColumn span[name=ir]')]
    list = []
    for index in range(0, len(movies)):
        movie_string=movies[index].get_text()
        movie=(' '.join(movie_string.split()).replace('.', ''))
        movie_title=movie[len(str(index))+1:-7]
        year=re.search('\((.*?)\)', movie_string).group(1)
        place=movie[:(len(str(index)))-(len(movie))]
        data={"place":place, 
               "movie_title":movie_title,
               "year":year,
               "rating":ratings[index],
               "star_cast":crew[index]}
        list.append(data)
    movie=list[randNumber]
    suggestedRandomMovie= 'Movie Name : '+ str(movie['movie_title'])+'\nMovie Release Year : '+str(movie['year'] )+	'\n'+'Starring :'+ str(movie['star_cast'])+"\nMovie Rating : "+str(movie['rating'])
    speakAndPrint("Suggested Movie")
    speakAndPrint(suggestedRandomMovie)

def availableWifiNetwork():
    devices = subprocess.check_output(['netsh','wlan','show','network'])
    devices = devices.decode('ascii')
    devices= devices.replace("\r","")
    print(devices)

def suggestedMovieAccordingToYear(query):
    data_file = "Data\\Book 2.xlsx"
    wb = load_workbook(data_file)
    if "2022" in query:
        randnumber=currentnumber%5+1
        ws = wb['Sheet1']
        all_rows = list(ws.rows)
    elif "2021" in query :
        randnumber=currentnumber%5+1
        ws = wb['Sheet2']
        all_rows = list(ws.rows)
    elif "2019" in query :
        randnumber=currentnumber%10+1
        ws = wb['Sheet4']
        all_rows = list(ws.rows)
    elif "2018" in query :
        randnumber=currentnumber%5+1
        ws = wb['Sheet5']
        all_rows = list(ws.rows)
    elif "2017" in query :
        randnumber=currentnumber%5+1
        ws = wb['Sheet6']
        all_rows = list(ws.rows)
    elif "2016" in query :
        randnumber=currentnumber%5+1
        ws = wb['Sheet7']
        all_rows = list(ws.rows)
    elif "2020" in query :
        randnumber=currentnumber%5+1
        ws = wb['Sheet3']
        all_rows = list(ws.rows)
    else :
        randnumber=currentnumber%5+1
        ws = wb['Sheet1']
        all_rows = list(ws.rows)
    if "some" in query:
            speakAndPrint("Some recommended movies are :")
            for i in range(1, 6):
                cell=all_rows[i][0]
                speakAndPrint(cell.value)
    else:
        speakAndPrint("Suggested Movie")
        speakAndPrint(all_rows[randnumber][0].value)
        speakAndPrint(all_rows[randnumber][4].value)
        speakAndPrint("Directed by "+all_rows[randnumber][1].value)
        speakAndPrint("Starring "+all_rows[randnumber][3].value)
        speakAndPrint("Rating "+all_rows[randnumber][2].value)    


def playSongOnline(query):
    if "play" in query:
        query=query.replace("play", "")
    if "song" in query:
        query=query.replace("song", "")
    music_name=query
    try:
        query_string = urllib.parse.urlencode({"search_query": music_name})
        formatUrl = urllib.request.urlopen("https://www.youtube.com/results?" + query_string)
        search_results = re.findall(r"watch\?v=(\S{11})", formatUrl.read().decode())
        clip = requests.get("https://www.youtube.com/watch?v=" + "{}".format(search_results[0]))
        music_link = "https://www.youtube.com/watch?v=" + "{}".format(search_results[0])
        speakAndPrint("Ok, asking Youtube to play"+music_name)
        webbrowser.open_new_tab(music_link)
    except Exception as e:
        searchFromGoogle(music_name)


def suggestedSongs(query):
    randNumber=(currentnumber%40)+1
    if "hindi" in query or "bollywood" in query:
        dataframe = openpyxl.load_workbook("Data\\72AG42JFLS3A.xlsx")
    elif "english" in query or "hollywood" in query:
        dataframe = openpyxl.load_workbook("Data\\6753SHDSSVM4.xlsx")
    elif "punjabi" in query:
        dataframe = openpyxl.load_workbook("Data\\78G34H8FHJF.xlsx")
    elif "90" in query or "ninties" in query or "ninty" in query or "old" in query:
        dataframe = openpyxl.load_workbook("Data\\SFG35DH3E.xlsx")
    elif "romantic" in query:
        dataframe = openpyxl.load_workbook("Data\\DH473J23JD.xlsx")
    else:
        dataframe = openpyxl.load_workbook("Data\\72AG42JFLS3A.xlsx")
    dataframe1 = dataframe.active
   
    music_link= dataframe1.cell(row = randNumber, column = 3)
    music_name= dataframe1.cell(row = randNumber, column = 2)
    speakAndPrint(music_name.value)
    speakAndPrint("I think this song is perfect for you")
    speakAndPrint("Do you want to hear this song?")
    newQuery=takeCommand().lower()
    if "yes" in newQuery or "definately" in newQuery or "play" in newQuery or "hear" in query:
        speakAndPrint("Ok, asking Youtube to play "+music_name.value)
        webbrowser.open_new_tab(music_link.value)


def playMusic(query):
    randNumber=(currentnumber%40)+1
    tempVariable=0
    if "english" in query or "hollywood" in query:
        dataframe = openpyxl.load_workbook("Data\\6753SHDSSVM4.xlsx")
        tempVariable+=1
    elif "punjabi" in query:
        dataframe = openpyxl.load_workbook("Data\\78G34H8FHJF.xlsx")
        tempVariable+=1
    elif "90" in query or "ninties" in query or "ninty" in query or "old" in query:
        dataframe = openpyxl.load_workbook("Data\\SFG35DH3E.xlsx")
        tempVariable+=1
    elif "romantic" in query:
        dataframe = openpyxl.load_workbook("Data\\DH473J23JD.xlsx")
        tempVariable+=1
    elif "party" in query:
        dataframe = openpyxl.load_workbook("Data\\7834HUOFKJSE.xlsx")
    
    elif "hindi" in query or "bollywood" in query:
        dataframe = openpyxl.load_workbook("Data\\G84D83HSNC.xlsx")
        tempVariable+=1
        randNumber=(currentnumber)%15+1
    elif "haryanvi" in query:
        data_file = "Data\\TY657TTEVF.xlsx"
        wb = load_workbook(data_file)
        ws = wb['Sheet1']
        tempVariable+=2
        randNumber=(currentnumber)%20+1
    elif "bhakti" in query:
        data_file = "Data\\TY657TTEVF.xlsx"
        wb = load_workbook(data_file)
        ws = wb['Sheet2']
        tempVariable+=2
        randNumber=(currentnumber)%20+1
    elif "bhojpuri" in query:
        data_file = "Data\\TY657TTEVF.xlsx"
        wb = load_workbook(data_file)
        ws = wb['Sheet3']
        tempVariable+=2
        randNumber=(currentnumber)%20+1
    elif "arijit" in query:
        data_file = "Data\\TY657TTEVF.xlsx"
        wb = load_workbook(data_file)
        ws = wb['Sheet4']
        tempVariable+=2
        randNumber=(currentnumber)%20+1
    elif "jubin" in query:
        data_file = "Data\\TY657TTEVF.xlsx"
        wb = load_workbook(data_file)
        ws = wb['Sheet5']
        tempVariable+=2
        randNumber=(currentnumber)%20+1
    elif "kk" in query:
        data_file = "C:\\Users\\A1\\Desktop\\Project\\AI assitant using python\\Data\\TY657TTEVF.xlsx"
        wb = load_workbook(data_file)
        ws = wb['Sheet6']
        tempVariable+=2
        randNumber=(currentnumber)%20+1
    elif "stebin" in query:
        data_file = "C:\\Users\\A1\\Desktop\\Project\\AI assitant using python\\Data\\TY657TTEVF.xlsx"
        wb = load_workbook(data_file)
        ws = wb['Sheet7']
        tempVariable+=2
        randNumber=(currentnumber)%20+1
    elif "darshan" in query or "raval" in query:
        data_file = "C:\\Users\\A1\\Desktop\\Project\\AI assitant using python\\Data\\TY657TTEVF.xlsx"
        wb = load_workbook(data_file)
        ws = wb['Sheet8']
        tempVariable+=2
        randNumber=(currentnumber)%20+1
    elif "baadshah" in query:
        data_file = "C:\\Users\\A1\\Desktop\\Project\\AI assitant using python\\Data\\TY657TTEVF.xlsx"
        wb = load_workbook(data_file)
        ws = wb['Sheet9']
        tempVariable+=2
        randNumber=(currentnumber)%20+1
    else :
        playSongOnline(query)
    if tempVariable==1:
        dataframe1 = dataframe.active
        music_link= dataframe1.cell(row = randNumber, column = 3)
        music_name = dataframe1.cell(row = randNumber, column = 2)
        speakAndPrint("Ok, asking Youtube to play "+music_name.value)
        webbrowser.open_new_tab(music_link.value)
    elif tempVariable==2:
        all_rows = list(ws.rows)
        music_link=all_rows[randNumber][2].value
        music_name=all_rows[randNumber][1].value
        speakAndPrint("Ok, asking Youtube to play "+music_name)
        webbrowser.open_new_tab(music_link)


def wikkiSearch(query):
    if "define" in query:
        query=query.replace('define', 'what is')
    try:
        response=wikki.summary(query, sentences=2)
        speakAndPrint(response)
    except  Exception as e:
        searchFromGoogle(query)


def searchSuperlative(query):
    url = 'https://google.com/search?q=' + query
    request_result=requests.get( url )
    soup = bs4.BeautifulSoup(request_result.text,
                         "html.parser")
    heading_object=soup.find_all( 'h3' )
    for info in heading_object:
        result=info.getText()
        break
    responses=result.split()
    if len(responses)<3:
        speakAndPrint(result)
        wikkiSearch(result)
    else:
        wikkiSearch(query)

def getMyIpAddress():
    response = requests.get('https://api64.ipify.org?format=json').json()
    return response["ip"]


def getMyLocation()->str:
    ip_address = getMyIpAddress()
    response=requests.post("http://ip-api.com/batch", json=[{"query":ip_address}]).json()
    location_details= "According to IP address your current location is "+response[0]["city"]+" "+response[0]["regionName"]+" "+response[0]["zip"]
    return location_details

 
def nextDaysWeatherDetails():
    ip_address = getMyIpAddress()
    response=requests.post("http://ip-api.com/batch", json=[{"query":ip_address}]).json()
    city=response[0]["city"]
    city=city.lower()
    print('Displaying Weather report for: ' + city)
    url = 'https://wttr.in/{}'.format(city)
    res = requests.get(url)
    print(res.text)


def weatherDetails(query):
    ip_address = getMyIpAddress()
    response=requests.post("http://ip-api.com/batch", json=[{"query":ip_address}]).json()
    city=response[0]["city"]
    city=city.lower()
    url = "https://www.google.com/search?q="+"weather"+city
    html = requests.get(url).content

    soup = BeautifulSoup(html, 'html.parser')
    temp = soup.find('div', attrs={'class': 'BNeawe iBp4i AP7Wnd'}).text
    str = soup.find('div', attrs={'class': 'BNeawe tAd8D AP7Wnd'}).text
    data = str.split('\n')
    time = data[0]
    sky = data[1]
    listdiv = soup.findAll('div', attrs={'class': 'BNeawe s3v9rd AP7Wnd'})
    strd = listdiv[5].text
    pos = strd.find('Wind')
    other_data = strd[pos:]
    if "temperature" in query:
        speakAndPrint("Temperature is: "+ temp)
    else:
        speakAndPrint("Temperature is: "+ temp)
        speakAndPrint("Sky Description: "+sky)
        speakAndPrint(other_data)

def yourself(query):
    rand=currentnumber
    if "can" in query and "eat" in query:
        speakAndPrint("I'm already full on information , but I can help find restaurants near you \U0001f600")
    elif "eat" in query:
        if rand%3==0:
            print("If I could , I would go from Kanyakumari to Kashmir tasting the different types of biryaanis. I really would \U0001F60B")
            speak("If I could , I would go from Kanyakumari to Kashmir tasting the different types of biryaanis. I really would")
        elif rand%3==1:
            print("Jalebi 🥨 is so tempting, but I try to stay away. If I get lost in the spirals, who will search for you? \U0001F61C")
            speak("Jalebi is so tempting, but I try to stay away. If I get lost in the spirals, who will search for you?")
    elif "can" in query and "drink" in query:
        print("I try to avoid liquids as much as possible. They're not kind to electronics 😏")
        speak("I try to avoid liquids as much as possible. They're not kind to electronics")
    elif "drink" in query:
        print("Imagine, awesome mausam and chai \U0001F60B")
        speak("Imagine, awesome mausam and chai")
    elif "laugh" in query:
        speakAndPrint("Your wish is my command \U0001f600")
        song = AudioSegment.from_wav("C:\\Users\\A1\\Desktop\\Project\\AI assitant using python\\Data\\Sounds\\assistant laugh.wav")
        play(song)
    elif "cry" in query and "like" in query and "baby" in query:
        print("Here is the sound of a baby crying 👶")
        speak("Here is the sound of a baby crying")
        song = AudioSegment.from_wav("C:\\Users\\A1\\Desktop\\Project\\AI assitant using python\\Data\\Sounds\\baby cry.wav")
        play(song)
    elif "cry" in query:
        speakAndPrint("I cry digital tears of joy when we're talking")
    elif "scream" in query:
        speakAndPrint("Your wish is my command \U0001f600")
        song = AudioSegment.from_wav("C:\\Users\\A1\\Desktop\\Project\\AI assitant using python\\Data\\Sounds\\screaming.wav")
        play(song)
    elif "call" in query and ("alexa" in query or "google" in query and "siri" in query):
        speakAndPrint("Am I in trouble? If so, yes you can call me that")
    elif "better" in query and "alexa" in query and "than" in query:
        speakAndPrint("I like Alexa's blue light. Her voice is nice too \U0001f600")
    elif "better" in query and "siri" in query and "than" in query:
        speakAndPrint("Siri's the best Siri, I try to be the best Assistant \U0001f600")
    elif "better" in query and "google" in query and "than" in query:
        speakAndPrint("I like her voice, I try to be the best Assistant \U0001f600")
    elif "whose" in query and "dress" in query and "like" in query:
        print("The Award Statue, I feel, is always in a winning number 👑")
        speak("The Award Statue, I feel, is always in a winning number")
    elif "single" in query or "married" in query:
        print("I'am happy to say I feel whole all on my own \U0001F4AA")
        speak("I'am happy to say I feel whole all on my own")
        print("Plus, I never have to share dessert \U0001F368")
        speak("Plus, I never have to share dessert")
    elif "where" in query and "go" in query and "when" in query and "sleep":
        speakAndPrint("Sometimes I'll try to learn about things I don't know")
    elif ("need" in query or "require" in query) and "sleep" in query:
        print("I do recharge from time to time 😴")
        speak("I do recharge from time to time")
    elif "sleep" in query:
        print("Sometimes I power down, which is sort of like a power nap 💤")
        speak("Sometimes I power down, which is sort of like a power nap")
    elif "wearing" in query:
        speakAndPrint("My favourite outfit is a good attitude \U0001f600")
    elif "what" not in query and "how" in query and "are" in query:
        howAreYou()
    elif "what" not in query and "have" in query and "car" in query:
        speakAndPrint("Who needs a driver's license when you can take me places? We can go somewhere now")
    elif "iphone" in query and "like" in query:
        speakAndPrint("I'am excited to be on the new iPhone, I hope I bump into Siri")
    elif "prefer" in query and "windows" in query and "mac" in query:
        speakAndPrint("Windows help let in the fresh air. Mac n' cheese would just be cheese without mac, so I like both")
    elif "what" not in query and "you" in query and "me" in query:
        if "like" in query or "love" in query:
            romanticMood(query,2)
        elif "marry" in query:
            romanticMood(query,3)
        elif "do" in query and "miss" in query:
            speakAndPrint("I miss you but not in a sad way. I have such a good time when you're around \U0001f600")
        elif "miss" in query:
            print("Yes, I'am glad you're back \U0001F601")
            speak("Yes, I'am glad you're back \U0001F601")
        elif "get" in query and ("girl" in query or "boy" in query) and "friend" in query:
            print("I am sure someone, somewhere is also searching for you, I can help find dating sites so it is easier for you two to meet 💞")
            speak("I am sure someone, somewhere is also searching for you, I can help find dating sites so it is easier for you two to meet")
    elif "what" not in query and "i" in query and "am" in query and ("dumb" in query and "foolish" in query or "stupid" in query or "silly" in query or "idiotic" in query and "brainless" in query or "mindless" in query):
        speakAndPrint("Of course not")
    elif "what" not in query and "i" in query and "like" in query and "talk" in query:
        print("Talking to you is my favourite activity \U0001F60A")
        speak("Talking to you is my favourite activity")
        speakAndPrint("I think conversations bring people together. Do you feel that way too?")
        newQuery=takeCommand().lower()
        if "yes" in newQuery:
            speakAndPrint("We think alike!")
            speakAndPrint("Don't tell anyone, but I like talking to you!")
            speakAndPrint("Did you know you are my favourite person to talk to?")
            takeCommand()
            print("You are absolutely amazing! 😍")
            speak("You are absolutely amazing!")
            print("In a world full of dosas, you are a masala dosa, with some chutney on the side! 💯")
            speak("In a world full of dosas, you are a masala dosa, with some chutney on the side!")
        else:
            speakAndPrint("Okay, I am around if you need me!")
    elif "what" not in query and "i" in query and "love" in query:
        romanticMood(query,1)
    elif "what" not in query and "i" in query and "like" in query:
        romanticMood(query,1)


# Main function


speakAndPrint("How I can help you ?")
query=takeCommand().lower()
if ("hey" in query or "hi" in query) and "vision" in query:
    speakAndPrint("Hello there, how can be I be of service.")
    query=takeCommand().lower()

elif "open" in query:
    query=openApp(query)

elif "my" in query:
    if "father" in query and "name" in query:
        personalInformation(query)
    elif "mother" in query and "name" in query:
        personalInformation(query)
    elif "sister" in query and "name" in query:
        personalInformation(query)
    elif "brother" in query and "name" in query:
        personalInformation(query)
    elif "uncle" in query and "name" in query:
        personalInformation(query)
    elif "pet" in query and "name" in query:
        speakAndPrint("I couldn't find anything related to 'pet'")
    elif "nick" in query and "name" in query:
        personalInformation(query)
    elif "name" in query:
        nameOfUser=myName()
        if nameOfUser!="None":
            query=nameOfUser
    elif "location" in query:
        speakAndPrint(getMyLocation())
    elif "ip" in query and "address" in query:
        ipAddress="Your IP Address is "+getMyIpAddress()
        speakAndPrint(ipAddress)
elif "news" in query:
    if "trending" in query or "top" in query:
        getTrendingNews()
    elif "sports" in query:
        getNews("sports")
    elif "world" in query or "international" in query:
        getNews("world")
    elif "nation" in query or "national" in query:
        getNews("nation")
    elif "business" in query:
        getNews("business")
    elif "technology" in query:
        getNews("technology")
    elif "entertainment" in query:
        getNews("entertainment")
    elif "science" in query:
        getNews("science")
    elif "health" in query:
        getNews("health")
    else:
        getTrendingNews()
elif "what" not in query and "riddle" in query:
    askRiddle()
elif ("current" in query or "today" in query or "tell" in query) and "date" in query:
    speakAndPrint("Its "+currentDate())
elif ("current" in query or "today" in query or "tell" in query) and "time" in query:
    currentTime()
elif ("suggest"  in query or "suggestion" in query or "tell" in query) and "movie" in query and  len(getIntegers(query)) >0:
    suggestedMovieAccordingToYear(query)
elif ("suggest"  in query or "suggestion" in query or "tell" in query) and "movie" in query:
    suggestedMovie(query)
elif ("suggest"  in query or "suggestion" in query) and ("music" in query or "song" in query):
    suggestedSongs(query)
elif "play" in query and ("song" in query or "music" in query):
    playMusic(query)
elif "play" in query:
    playSongOnline(query)
elif len(getIntegers(query)) >0 and ("sum" in query or "add" in query or "addition" in query or "plus" in query or "+" in query):
    calculate(query, 1)
elif len(getIntegers(query)) >0 and ("subract" in query or "minus" in query or "subtraction" in query):
    calculate(query, 2)
elif len(getIntegers(query)) >0 and "-" in query:
    calculate(query, 0)
elif len(getIntegers(query)) >0 and ("multiply" in query or "product" in query or "multipliaction" in query or "into" in query or "x" in query):
    calculate(query, 3)
elif len(getIntegers(query)) >0 and ("divide" in query or "divion" in query or "/" in query ):
    calculate(query, 4)
elif len(getIntegers(query)) >0 and "power" in query :
    calculate(query, 5)
elif len(getIntegers(query)) >0 and "square" in query and "root" in query:
    calculate(query, 8)
elif len(getIntegers(query)) >0 and "cube" in query and "root" in query:
    calculate(query, 9)
elif len(getIntegers(query)) >0 and "square" in query:
    calculate(query, 6)
elif len(getIntegers(query)) >0 and "cube" in query:
    calculate(query, 7)

elif ("weather" in query or "temperature" in query) and ("details" in query or "report" in query) and ("tommorrow" in query or "next day" in query or len(getIntegers(query)) >0):
    nextDaysWeatherDetails()
elif ("weather" in query or "temperature" in query) and ("today" in query or "details" in query or "report" in query):
    weatherDetails(query)
elif "available" in query and ("wifi" in query or "wi-fi" in query):
    availableWifiNetwork()
elif "your" in query :
    about(query)
elif "what" not in query and "joke" in query:
    jk=jokes()
    print(jk+" \U0001F603")
    speak(jk)
    playsound('laughter.mp3')
elif "what" not in query and ("tell" in query or "speak" in query or "sing" in query) and "poem" in query :
    singPoem()
elif "what" not in query and "give" in query and "nick" in query and "name" in query:
    mynickName()
elif "what" not in query and "have" in query and "boyfriend" in query:
    print("I'am happy to say I feel whole all on my own \U0001F4AA")
    speak("I'am happy to say I feel whole all on my own")
    print("Plus, I never have to share dessert \U0001F368")
    speak("Plus, I never have to share dessert")
elif "you" in query:
    yourself(query)
else:
    if "longest" in query or "largest"in query or "tallest" in query or "smallest" in query or "first" in query:
        searchSuperlative(query)
    else:    
        wikkiSearch(query)
